"""主窗口 - PDF 批量添加文字工具的主界面"""
import os
import sys
import subprocess
import tempfile
import re
import difflib
import time
from datetime import datetime

import fitz
import openpyxl

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QProgressBar, QSpinBox,
    QComboBox, QCheckBox, QHeaderView, QPlainTextEdit,
    QDialog, QColorDialog, QDoubleSpinBox,
    QFrame, QGridLayout, QScrollArea, QSizePolicy, QMenu,
    QStyledItemDelegate, QStyle, QInputDialog
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QPointF, QTimer, QSettings
from PyQt6.QtGui import QFont, QColor, QCursor, QPen, QShortcut, QKeySequence

from .config import (
    COLORS, APP_NAME, APP_VERSION, POSITIONS, DEFAULT_FONT_SIZE,
    DEFAULT_TEXT_COLOR, DEFAULT_POSITION, FITZ_FONT_MAP, CJK_FONT_RESOURCE_NAME,
    CHECKPOINT_DIR, CHECKPOINT_FILE
)
from .logger import diag_log, diag_flush
from .utils.pages import parse_page_range
from .utils.fonts import find_cjk_font
from .utils.checkpoint import save_checkpoint, load_checkpoint, clear_checkpoint
from .utils.history import load_watermark_history, save_watermark_history, smart_recommend_text
from .utils.notify import send_system_notification
from .utils.auto_update import check_for_update, download_and_update, GITHUB_OWNER, GITHUB_REPO
from .utils.templates import load_watermark_templates, save_watermark_templates
from .utils.ensure import create_app_icon
from .widgets import TextColumnDelegate, SmartSuggestWidget
from .dialogs import (
    BatchEditDialog, SettingsDialog, PdfPreviewDialog,
    WatermarkTemplateDialog, PreflightDialog, PdfToolsDialog, MultiTextDialog, PDFAdvancedDialog
)
from .pdf.workers import ProcessWorker, PageCountWorker

# 目录文件列表缓存
_DIR_FILELIST_CACHE = {}


class MainWindow(QMainWindow):
    # 全局样式表
    GLOBAL_CSS = f"""
    QMainWindow {{
        background-color: {COLORS['bg']};
    }}
    /* 滚动条 */
    QScrollBar:vertical {{
        background-color: transparent;
        width: 6px;
        margin: 0px;
    }}
    QScrollBar::handle:vertical {{
        background-color: {COLORS['border']};
        border-radius: 3px;
        min-height: 40px;
    }}
    QScrollBar::handle:vertical:hover {{
        background-color: {COLORS['text_muted']};
    }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
        height: 0px;
    }}
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
        background: transparent;
    }}
    QScrollBar:horizontal {{
        background-color: transparent;
        height: 6px;
        margin: 0px;
    }}
    QScrollBar::handle:horizontal {{
        background-color: {COLORS['border']};
        border-radius: 3px;
        min-width: 40px;
    }}
    QScrollBar::handle:horizontal:hover {{
        background-color: {COLORS['text_muted']};
    }}
    QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
        width: 0px;
    }}
    """

    # 卡片容器样式
    CARD_CSS = f"""
    background-color: {COLORS['card']};
    border: 1px solid {COLORS['border']};
    border-radius: 14px;
    """

    # 输入框通用样式
    INPUT_CSS = f"""
    QLineEdit {{
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        padding: 10px 14px;
        background-color: white;
        color: {COLORS['text']};
        font-size: 13px;
        selection-background-color: {COLORS['primary_light']};
        selection-color: {COLORS['text']};
    }}
    QLineEdit:hover {{
        border-color: {COLORS['primary']};
    }}
    QLineEdit:focus {{
        border: 1.5px solid {COLORS['primary']};
        padding: 9px 13px;
    }}
    QLineEdit[readOnly="true"] {{
        background-color: {COLORS['border_light']};
        color: {COLORS['text_secondary']};
    }}
    """

    # 主要按钮样式
    PRIMARY_BTN_CSS = f"""
    QPushButton {{
        background-color: {COLORS['primary']};
        color: white;
        border: none;
        border-radius: 8px;
        padding: 11px 26px;
        font-weight: 600;
        font-size: 14px;
    }}
    QPushButton:hover {{
        background-color: {COLORS['primary_hover']};
    }}
    QPushButton:pressed {{
        background-color: {COLORS['primary_hover']};
    }}
    QPushButton:disabled {{
        background-color: {COLORS['border']};
        color: {COLORS['text_muted']};
    }}
    """

    # 次要按钮样式
    SECONDARY_BTN_CSS = f"""
    QPushButton {{
        background-color: {COLORS['primary_light']};
        color: {COLORS['primary']};
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        padding: 10px 18px;
        font-weight: 600;
        font-size: 13px;
    }}
    QPushButton:hover {{
        background-color: {COLORS['primary']};
        color: white;
        border-color: {COLORS['primary']};
    }}
    """

    # 成功按钮样式
    SUCCESS_BTN_CSS = f"""
    QPushButton {{
        background-color: {COLORS['accent']};
        color: white;
        border: none;
        border-radius: 10px;
        padding: 14px 28px;
        font-weight: 700;
        font-size: 15px;
    }}
    QPushButton:hover {{
        background-color: {COLORS['accent_hover']};
    }}
    QPushButton:disabled {{
        background-color: {COLORS['border']};
        color: {COLORS['text_muted']};
    }}
    """

    # 警告按钮样式
    WARNING_BTN_CSS = f"""
    QPushButton {{
        background-color: {COLORS['warning_light']};
        color: {COLORS['warning']};
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        padding: 10px 18px;
        font-weight: 600;
        font-size: 13px;
    }}
    QPushButton:hover {{
        background-color: {COLORS['warning']};
        color: white;
        border-color: {COLORS['warning']};
    }}
    """

    # 危险按钮样式
    DANGER_BTN_CSS = f"""
    QPushButton {{
        background-color: {COLORS['danger_light']};
        color: {COLORS['danger']};
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        padding: 9px 16px;
        font-weight: 600;
        font-size: 12px;
    }}
    QPushButton:hover {{
        background-color: {COLORS['danger']};
        color: white;
        border-color: {COLORS['danger']};
    }}
    """

    # 表格样式
    TABLE_CSS = f"""
    QTableWidget {{
        border: none;
        background-color: white;
        gridline-color: {COLORS['border_light']};
        font-size: 13px;
        selection-background-color: {COLORS['primary']};
        selection-color: white;
        alternate-background-color: {COLORS['primary_ultra_light']};
    }}
    QTableWidget::item {{
        padding: 6px 4px;
        border-bottom: 1px solid {COLORS['border_light']};
    }}
    QTableWidget::item:selected {{
        background-color: {COLORS['primary']};
        color: white;
    }}
    QHeaderView::section {{
        background-color: {COLORS['bg']};
        padding: 10px 8px;
        border: none;
        border-bottom: 1.5px solid {COLORS['primary']};
        font-weight: 600;
        color: {COLORS['text_secondary']};
        font-size: 12px;
    }}
    QLineEdit {{
        padding: 4px;
        border: 1px solid {COLORS['primary']};
        border-radius: 2px;
        background: white;
        font-size: 13px;
    }}
    """

    # 日志区域样式
    LOG_CSS = f"""
    QPlainTextEdit {{
        background-color: #2C3E50;
        color: #B0C4D8;
        font-family: 'Consolas', 'Monaco', monospace;
        font-size: 12px;
        border-radius: 8px;
        padding: 14px;
        border: 1px solid {COLORS['border']};
    }}
    """

    # 进度条样式
    PROGRESS_CSS = f"""
    QProgressBar {{
        border: none;
        border-radius: 10px;
        text-align: center;
        background-color: {COLORS['border_light']};
        color: {COLORS['text_secondary']};
        font-size: 11px;
        font-weight: 600;
    }}
    QProgressBar::chunk {{
        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
            stop:0 {COLORS['header_bg']}, stop:1 {COLORS['header_bg_end']});
        border-radius: 10px;
    }}
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION}")
        self.setMinimumSize(1150, 750)
        self.resize(1250, 800)
        self.setWindowIcon(create_app_icon())
        self.setAcceptDrops(True)

        self.settings = QSettings("PDFBatchTool", "AddText")
        self.tasks = []
        self.worker = None
        self.text_settings = {}

        self.setup_ui()
        self.load_settings()
        self.apply_theme()
        self._setup_shortcuts()

        # 后台检查更新
        QTimer.singleShot(1000, self._check_update_on_startup)

        # 检查是否有中断的会话，提供恢复选项
        QTimer.singleShot(500, self._check_for_resume)

    def _make_card(self, parent_layout, spacing=18, margins=(24, 22, 24, 22)):
        """创建卡片容器并返回其布局"""
        card = QWidget()
        card.setStyleSheet(self.CARD_CSS)
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        layout = QVBoxLayout(card)
        layout.setSpacing(spacing)
        layout.setContentsMargins(*margins)
        parent_layout.addWidget(card)
        return card, layout

    def _section_title(self, text, subtitle=None):
        """创建区域标题"""
        container = QHBoxLayout()
        container.setSpacing(12)

        title = QLabel(text)
        title.setFont(QFont("", 15, QFont.Weight.Bold))
        title.setStyleSheet(f"color:{COLORS['text']};border:none;background:transparent;")
        container.addWidget(title)

        if subtitle:
            sub = QLabel(subtitle)
            sub.setStyleSheet(f"color:{COLORS['text_muted']};font-size:12px;border:none;background:transparent;")
            container.addWidget(sub)

        container.addStretch()
        return container, title

    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # === 顶部标题栏 ===
        header = QWidget()
        header.setFixedHeight(72)
        header.setStyleSheet(f"""
            QWidget#headerBar {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['header_bg']}, stop:1 {COLORS['header_bg_end']});
                border-bottom: none;
            }}
        """)
        header.setObjectName("headerBar")
        hl = QHBoxLayout(header)
        hl.setContentsMargins(32, 0, 32, 0)
        hl.setSpacing(14)

        icon_label = QLabel("")
        icon_label.setPixmap(create_app_icon().pixmap(36, 36))
        hl.addWidget(icon_label)

        tl = QVBoxLayout()
        tl.setSpacing(2)
        t = QLabel(APP_NAME)
        t.setFont(QFont("", 16, QFont.Weight.Bold))
        t.setStyleSheet("color:white;border:none;background:transparent;")
        tl.addWidget(t)
        s = QLabel("导入 Excel 批量匹配 PDF，自动添加指定文字")
        s.setStyleSheet("color:rgba(255,255,255,0.75);font-size:11px;border:none;background:transparent;")
        tl.addWidget(s)
        hl.addLayout(tl)
        hl.addStretch()

        ver = QLabel(f" v{APP_VERSION} ")
        ver.setStyleSheet("color:rgba(255,255,255,0.85);font-size:11px;padding:4px 14px;"
                         "background-color:rgba(255,255,255,0.15);border-radius:20px;border:none;")
        hl.addWidget(ver)

        main_layout.addWidget(header)

        # === 可滚动主内容区 ===
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet(f"QScrollArea{{border:none;background-color:{COLORS['bg']};}}")

        content = QWidget()
        content.setStyleSheet(f"background-color:{COLORS['bg']};")
        cl = QVBoxLayout(content)
        cl.setSpacing(18)
        cl.setContentsMargins(28, 20, 28, 24)

        # ---- 文件导入区 ----
        _, il = self._make_card(cl, spacing=16, margins=(24, 20, 24, 20))
        hdr, _ = self._section_title("文件导入", "选择数据源和目标文件夹")
        il.addLayout(hdr)

        # Excel 行
        erow = QHBoxLayout()
        erow.setSpacing(10)
        el = QLabel("Excel 文件")
        el.setFixedWidth(80)
        el.setStyleSheet(f"color:{COLORS['text_secondary']};font-size:13px;font-weight:500;border:none;background:transparent;")
        erow.addWidget(el)

        self.excel_path = QLineEdit()
        self.excel_path.setPlaceholderText("点击浏览选择包含任务列表的 Excel 文件...")
        self.excel_path.setReadOnly(True)
        self.excel_path.setMinimumHeight(40)
        self.excel_path.setStyleSheet(self.INPUT_CSS)
        self.excel_path.setToolTip("导入包含 序号/文件名/添加文字 的 Excel 文件 (Ctrl+O)")
        erow.addWidget(self.excel_path, stretch=1)

        eb = QPushButton("浏览")
        eb.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        eb.setFixedSize(72, 40)
        eb.setStyleSheet(self.SECONDARY_BTN_CSS)
        eb.clicked.connect(self.browse_excel)
        erow.addWidget(eb)

        tb = QPushButton("模板")
        tb.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        tb.setFixedSize(64, 40)
        tb.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['accent_light']};
                color: {COLORS['accent']};
                border: none;
                border-radius: 10px;
                font-weight: 600;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['accent']};
                color: white;
            }}
        """)
        tb.clicked.connect(self.download_template)
        erow.addWidget(tb)

        # 智能生成按钮
        ag_btn = QPushButton("智能生成")
        ag_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        ag_btn.setFixedSize(80, 40)
        ag_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['accent']};
                color: white;
                border: none;
                border-radius: 10px;
                font-weight: 600;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['accent_hover']};
            }}
        """)
        ag_btn.setToolTip("扫描 PDF 文件夹，自动为每个 PDF 生成任务行")
        ag_btn.clicked.connect(self.auto_generate_excel)
        erow.addWidget(ag_btn)

        il.addLayout(erow)

        # PDF 文件夹行
        prow = QHBoxLayout()
        prow.setSpacing(10)
        pl = QLabel("PDF 文件夹")
        pl.setFixedWidth(80)
        pl.setStyleSheet(f"color:{COLORS['text_secondary']};font-size:13px;font-weight:500;border:none;background:transparent;")
        prow.addWidget(pl)

        self.pdf_dir = QLineEdit()
        self.pdf_dir.setPlaceholderText("点击浏览选择存放 PDF 文件的文件夹...")
        self.pdf_dir.setReadOnly(True)
        self.pdf_dir.setMinimumHeight(40)
        self.pdf_dir.setStyleSheet(self.INPUT_CSS)
        self.pdf_dir.setToolTip("选择 PDF 文件所在的文件夹，加载 Excel 后可自动检测 (Ctrl+D)")
        prow.addWidget(self.pdf_dir, stretch=1)

        pb = QPushButton("浏览")
        pb.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        pb.setFixedSize(72, 40)
        pb.setStyleSheet(self.SECONDARY_BTN_CSS)
        pb.clicked.connect(self.browse_pdf_dir)
        prow.addWidget(pb)
        il.addLayout(prow)

        # 输出目录行
        orow = QHBoxLayout()
        orow.setSpacing(10)
        ol = QLabel("输出目录")
        ol.setFixedWidth(80)
        ol.setStyleSheet(f"color:{COLORS['text_secondary']};font-size:13px;font-weight:500;border:none;background:transparent;")
        orow.addWidget(ol)

        self.output_dir = QLineEdit()
        self.output_dir.setPlaceholderText("默认在 PDF 文件夹内创建 output 子文件夹")
        self.output_dir.setReadOnly(True)
        self.output_dir.setMinimumHeight(40)
        self.output_dir.setStyleSheet(self.INPUT_CSS)
        self.output_dir.setToolTip("处理后的 PDF 保存位置，默认在 PDF 文件夹下创建 output 子目录")
        orow.addWidget(self.output_dir, stretch=1)

        ob = QPushButton("浏览")
        ob.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        ob.setFixedSize(72, 40)
        ob.setStyleSheet(self.SECONDARY_BTN_CSS)
        ob.clicked.connect(self.browse_output_dir)
        orow.addWidget(ob)
        il.addLayout(orow)

        # 加载按钮行
        lrow = QHBoxLayout()
        lrow.addStretch()
        self.load_btn = QPushButton("加载 Excel 数据")
        self.load_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.load_btn.setFixedSize(180, 44)
        self.load_btn.setStyleSheet(self.PRIMARY_BTN_CSS)
        self.load_btn.clicked.connect(self.load_excel)
        self.load_btn.setToolTip("读取 Excel 文件中的任务数据并填充到下方表格")
        lrow.addWidget(self.load_btn)
        lrow.addStretch()
        il.addLayout(lrow)

        # ---- 中间区域: 样式 + 任务列表 ----
        mid = QHBoxLayout()
        mid.setSpacing(18)

        # 左侧面板
        left_card, ll = self._make_card(mid, spacing=16, margins=(22, 20, 22, 20))
        left_card.setFixedWidth(260)

        stitle_layout, _ = self._section_title("文字样式")
        ll.addLayout(stitle_layout)

        # 预览区
        prev = QWidget()
        prev.setFixedHeight(72)
        prev.setStyleSheet(f"""
            background-color:{COLORS['bg']};
            border-radius:12px;
            border:1.5px dashed {COLORS['border']};
        """)
        pl_ = QVBoxLayout(prev)
        pl_.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_text = QLabel("预览文字")
        self.preview_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_text.setFont(QFont("", 15))
        self.preview_text.setStyleSheet(f"color:{COLORS['primary']};border:none;background:transparent;")
        pl_.addWidget(self.preview_text)
        ll.addWidget(prev)

        self.style_detail = QLabel("14pt | 右下角")
        self.style_detail.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.style_detail.setStyleSheet(f"color:{COLORS['text_muted']};font-size:11px;border:none;background:transparent;")
        ll.addWidget(self.style_detail)

        sb = QPushButton("修改样式设置")
        sb.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        sb.setFixedHeight(38)
        sb.setStyleSheet(self.SECONDARY_BTN_CSS)
        sb.clicked.connect(self.open_settings)
        sb.setToolTip("设置文字字体、大小、颜色、位置、透明度等")
        ll.addWidget(sb)

        # 分隔
        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.HLine)
        sep1.setStyleSheet(f"background-color:{COLORS['border']};max-height:1px;")
        ll.addWidget(sep1)

        # 统计
        stats = QVBoxLayout()
        stats.setSpacing(6)
        self.stat_total = QLabel("总任务 0")
        self.stat_total.setStyleSheet(f"color:{COLORS['text_secondary']};font-size:12px;border:none;background:transparent;")
        stats.addWidget(self.stat_total)
        self.stat_ready = QLabel("就绪 0")
        self.stat_ready.setStyleSheet(f"color:{COLORS['accent']};font-size:12px;font-weight:600;border:none;background:transparent;")
        stats.addWidget(self.stat_ready)
        self.stat_missing = QLabel("未找到 0")
        self.stat_missing.setStyleSheet(f"color:{COLORS['danger']};font-size:12px;border:none;background:transparent;")
        stats.addWidget(self.stat_missing)
        ll.addLayout(stats)
        ll.addStretch()

        cb = QPushButton("清空任务列表")
        cb.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        cb.setFixedHeight(34)
        cb.setStyleSheet(self.DANGER_BTN_CSS)
        cb.clicked.connect(self.clear_tasks)
        ll.addWidget(cb)

        # 右侧: 任务列表
        _, rl = self._make_card(mid, spacing=14, margins=(22, 18, 22, 18))

        thdr = QHBoxLayout()
        tt, _ = self._section_title("任务列表")
        thdr.addLayout(tt)

        self.task_count = QLabel("0 个任务")
        self.task_count.setStyleSheet(
            f"color:{COLORS['text_muted']};font-size:11px;padding:4px 14px;"
            f"background-color:{COLORS['border_light']};border-radius:20px;border:none;"
        )
        thdr.addWidget(self.task_count)
        thdr.addStretch()

        # 搜索过滤框
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索文件名/文字...")
        self.search_input.setFixedWidth(180)
        self.search_input.setFixedHeight(28)
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                border: 1px solid {COLORS['border']};
                border-radius: 14px;
                padding: 2px 12px;
                font-size: 11px;
                background: {COLORS['primary_ultra_light']};
                color: {COLORS['text']};
            }}
            QLineEdit:focus {{
                border-color: {COLORS['primary']};
            }}
        """)
        self.search_input.textChanged.connect(self._on_search_changed)
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(200)
        self._search_timer.timeout.connect(self._do_filter_table)
        thdr.addWidget(self.search_input)

        # 批量编辑按钮
        batch_edit_btn = QPushButton("✏️ 批量编辑")
        batch_edit_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        batch_edit_btn.setFixedHeight(28)
        batch_edit_btn.setStyleSheet(f"""
            QPushButton {{ background:{COLORS['accent_light']}; color:{COLORS['accent']};
            border:1px solid {COLORS['border']}; border-radius:14px; padding:2px 14px;
            font-size:11px; font-weight:600; }}
            QPushButton:hover {{ background:{COLORS['accent']}; color:white; }}
        """)
        batch_edit_btn.clicked.connect(self._open_batch_edit)
        batch_edit_btn.setToolTip("批量编辑所有任务的文字内容（查找替换、前置追加、正则等）")
        thdr.addWidget(batch_edit_btn)

        rl.addLayout(thdr)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["序号", "文件名", "添加文字", "页码", "页数", "PDF 路径", "状态"])
        self.table.horizontalHeader().sectionClicked.connect(self._on_header_sort)
        self._sort_col = -1
        self._sort_order = Qt.SortOrder.AscendingOrder
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 50)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(3, 80)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(4, 55)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(6, 80)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.verticalHeader().setVisible(False)
        self.table.setMinimumHeight(200)
        self.table.setStyleSheet(self.TABLE_CSS)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_table_context_menu)
        self.table.doubleClicked.connect(lambda idx: self._on_table_double_click(idx.row(), idx.column()))
        self.table.cellChanged.connect(self._on_cell_changed)
        # 设置委托：仅"添加文字"列可编辑
        self.table.setItemDelegateForColumn(2, TextColumnDelegate(2, self))
        rl.addWidget(self.table, stretch=1)

        # 智能推荐面板
        self.suggest_widget = SmartSuggestWidget()
        self.suggest_widget.text_selected.connect(self._on_suggestion_selected)
        rl.addWidget(self.suggest_widget)

        # 输出命名规则
        naming_layout = QHBoxLayout()
        naming_layout.setSpacing(6)
        nl = QLabel("输出命名:")
        nl.setStyleSheet(f"color:{COLORS['text_muted']};font-size:11px;font-weight:500;border:none;background:transparent;")
        naming_layout.addWidget(nl)
        self.naming_template = QComboBox()
        self.naming_template.setEditable(True)
        self.naming_template.addItems([
            "{原文件名}.pdf",
            "{原文件名}_{文字}.pdf",
            "{原文件名}_{日期}.pdf",
            "{文字}_{原文件名}.pdf",
            "{序号:03d}_{原文件名}.pdf",
        ])
        self.naming_template.setCurrentText(self.settings.value("naming_template", "{原文件名}.pdf"))
        self.naming_template.setFixedWidth(240)
        self.naming_template.setFixedHeight(26)
        self.naming_template.setStyleSheet(f"""
            QComboBox {{ border:1px solid {COLORS['border']}; border-radius:4px;
            padding:2px 6px; font-size:11px; background:white; color:{COLORS['text']}; }}
            QComboBox:focus {{ border-color:{COLORS['primary']}; }}
        """)
        self.naming_template.setToolTip("输出文件命名模板\n{原文件名}=原文件名 {文字}=添加文字 {日期}=日期 {时间}=时间 {序号}=序号")
        naming_layout.addWidget(self.naming_template)
        naming_layout.addStretch()
        rl.addLayout(naming_layout)

        cl.addLayout(mid)

        # ---- 底部: 日志 + 控制 ----
        bottom = QHBoxLayout()
        bottom.setSpacing(18)

        # 日志
        _, lol = self._make_card(bottom, spacing=12, margins=(22, 18, 22, 18))
        log_hdr, _ = self._section_title("处理日志")
        lol.addLayout(log_hdr)

        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumBlockCount(300)
        self.log_text.setStyleSheet(self.LOG_CSS)
        self.log_text.setFixedHeight(130)
        lol.addWidget(self.log_text)

        # 控制面板
        ctrl_card, ctrl_l = self._make_card(bottom, spacing=14, margins=(22, 18, 22, 18))
        ctrl_card.setFixedWidth(260)

        ctrl_hdr, _ = self._section_title("操作控制")
        ctrl_l.addLayout(ctrl_hdr)

        self.progress = QProgressBar()
        self.progress.setTextVisible(True)
        self.progress.setFixedHeight(24)
        self.progress.setStyleSheet(self.PROGRESS_CSS)
        ctrl_l.addWidget(self.progress)

        self.status_label = QLabel(f"就绪 | {APP_NAME} v{APP_VERSION}")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.status_label.setStyleSheet(
            f"color:{COLORS['text_muted']};font-size:13px;font-weight:600;"
            f"border:1px solid {COLORS['border']};border-radius:8px;"
            f"padding:4px 16px;background:{COLORS['card']};"
        )
        self.status_label.setToolTip("点击检查更新 | GitHub Release")
        self.status_label.mousePressEvent = lambda e: self._check_update_now()
        ctrl_l.addWidget(self.status_label)
        ctrl_l.addStretch()

        self.start_btn = QPushButton("开始处理")
        self.start_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.start_btn.setFixedHeight(46)
        self.start_btn.setStyleSheet(self.SUCCESS_BTN_CSS)
        self.start_btn.clicked.connect(self.start_process)
        self.start_btn.setToolTip("开始批量添加文字到 PDF (Ctrl+Enter)")
        ctrl_l.addWidget(self.start_btn)

        # 一键智能处理按钮
        self.smart_btn = QPushButton("🤖 一键智能")
        self.smart_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.smart_btn.setFixedHeight(42)
        self.smart_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6366F1, stop:1 #8B5CF6);
                color: white; border: none; border-radius: 10px;
                padding: 10px 18px; font-weight: 700; font-size: 14px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4F46E5, stop:1 #7C3AED);
            }}
            QPushButton:disabled {{
                background-color: {COLORS['border']};
                color: {COLORS['text_muted']};
            }}
        """)
        self.smart_btn.clicked.connect(self.one_click_smart_process)
        self.smart_btn.setToolTip("智能全自动：扫描PDF→生成任务→添加文字→输出报告")
        ctrl_l.addWidget(self.smart_btn)

        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.stop_btn.setFixedHeight(38)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet(self.WARNING_BTN_CSS)
        self.stop_btn.clicked.connect(self.stop_process)
        self.stop_btn.setToolTip("停止当前处理")
        ctrl_l.addWidget(self.stop_btn)

        self.preview_btn = QPushButton("预览效果")
        self.preview_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.preview_btn.setFixedHeight(38)
        self.preview_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.preview_btn.clicked.connect(self.preview_pdf)
        self.preview_btn.setToolTip("预览第一个 PDF 的文字添加效果")
        ctrl_l.addWidget(self.preview_btn)

        self.reset_btn = QPushButton("重置")
        self.reset_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.reset_btn.setFixedHeight(38)
        self.reset_btn.setStyleSheet(self.DANGER_BTN_CSS)
        self.reset_btn.clicked.connect(self.reset_all)
        self.reset_btn.setToolTip("清空所有设置和任务列表")
        ctrl_l.addWidget(self.reset_btn)

        self.retry_btn = QPushButton("重试失败项")
        self.retry_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.retry_btn.setFixedHeight(34)
        self.retry_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.retry_btn.clicked.connect(self.retry_failed)
        self.retry_btn.setToolTip("仅重新处理状态为'失败'的任务")
        self.retry_btn.setVisible(False)
        ctrl_l.addWidget(self.retry_btn)

        # 生成报告按钮
        self.report_btn = QPushButton("📊 生成报告")
        self.report_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.report_btn.setFixedHeight(34)
        self.report_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.report_btn.clicked.connect(self.generate_report)
        self.report_btn.setToolTip("生成批量处理结果报告")
        self.report_btn.setVisible(False)
        ctrl_l.addWidget(self.report_btn)

        # 水印模板按钮
        self.template_btn = QPushButton("📋 水印模板")
        self.template_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.template_btn.setFixedHeight(34)
        self.template_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.template_btn.clicked.connect(self._open_template_manager)
        self.template_btn.setToolTip("管理/应用水印模板（红章合同、财务发票等）")
        ctrl_l.addWidget(self.template_btn)

        # 智能分类按钮
        self.classify_btn = QPushButton("📂 智能分类")
        self.classify_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.classify_btn.setFixedHeight(34)
        self.classify_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.classify_btn.clicked.connect(self._smart_classify_tasks)
        self.classify_btn.setToolTip("根据文件名自动分类并设置不同水印样式")
        ctrl_l.addWidget(self.classify_btn)

        # 批量多文字按钮
        self.multi_text_btn = QPushButton("📝 批量多文字")
        self.multi_text_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.multi_text_btn.setFixedHeight(34)
        self.multi_text_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.multi_text_btn.clicked.connect(self._open_multi_text)
        self.multi_text_btn.setToolTip("为同一 PDF 的不同位置添加不同文字，各文字可独立设置内容、颜色、大小、位置")
        ctrl_l.addWidget(self.multi_text_btn)

        # 一键还原按钮
        self.restore_btn = QPushButton("🔄 一键还原")
        self.restore_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.restore_btn.setFixedHeight(34)
        self.restore_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.restore_btn.clicked.connect(self._restore_from_backup)
        self.restore_btn.setToolTip("还原原始 PDF 文件（从备份目录）")
        ctrl_l.addWidget(self.restore_btn)

        # 智能同步按钮
        self.sync_btn = QPushButton("🔗 同步水印")
        self.sync_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.sync_btn.setFixedHeight(34)
        self.sync_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.sync_btn.clicked.connect(self._smart_sync_watermark)
        self.sync_btn.setToolTip("将选中任务的水印同步到文件名相似的任务")
        ctrl_l.addWidget(self.sync_btn)

        # 页面适应按钮
        self.adapt_btn = QPushButton("📐 页面适应")
        self.adapt_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.adapt_btn.setFixedHeight(34)
        self.adapt_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.adapt_btn.clicked.connect(self._adaptive_page_size)
        self.adapt_btn.setToolTip("检测 PDF 页面大小，自动推荐字号和位置")
        ctrl_l.addWidget(self.adapt_btn)

        # 智能预检按钮
        self.preflight_btn = QPushButton("🔍 智能预检")
        self.preflight_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.preflight_btn.setFixedHeight(34)
        self.preflight_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.preflight_btn.clicked.connect(lambda: self._run_preflight_check())
        self.preflight_btn.setToolTip("处理前扫描所有 PDF 状态（文件是否存在、加密、空文件等）")
        ctrl_l.addWidget(self.preflight_btn)

        # PDF 工具集按钮
        self.tools_btn = QPushButton("🔧 PDF 工具")
        self.tools_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.tools_btn.setFixedHeight(34)
        self.tools_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.tools_btn.clicked.connect(self._open_pdf_tools)
        self.tools_btn.setToolTip("添加页码/页脚，或从 PDF 提取文本")
        ctrl_l.addWidget(self.tools_btn)

        # PDF 增强工具按钮
        self.advanced_btn = QPushButton("⚡ 增强工具")
        self.advanced_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.advanced_btn.setFixedHeight(34)
        self.advanced_btn.setStyleSheet(self.SECONDARY_BTN_CSS)
        self.advanced_btn.clicked.connect(self._open_advanced_tools)
        self.advanced_btn.setToolTip("图片水印、页面旋转、PDF加密")
        ctrl_l.addWidget(self.advanced_btn)

        ctrl_l.addStretch()
        cl.addLayout(bottom)

        scroll.setWidget(content)
        main_layout.addWidget(scroll, stretch=1)

        # === 底部状态栏 ===
        footer = QWidget()
        footer.setFixedHeight(32)
        footer.setStyleSheet(f"background-color:{COLORS['card']};border-top:1px solid {COLORS['border']};")
        fl = QHBoxLayout(footer)
        fl.setContentsMargins(24, 0, 24, 0)

        dev_info = QLabel("开发者: 彭鹏  |  联系方式: 18551346387  |  Ctrl+O 导入  Ctrl+Enter 开始  Ctrl+E 导出  Del 删除选中")
        dev_info.setStyleSheet(f"color:{COLORS['text_muted']};font-size:10px;border:none;background:transparent;")
        fl.addWidget(dev_info)
        fl.addStretch()

        main_layout.addWidget(footer)

    def apply_theme(self):
        """应用全局样式表"""
        self.setStyleSheet(self.GLOBAL_CSS)

    def _setup_shortcuts(self):
        """注册键盘快捷键"""
        QShortcut(QKeySequence("Ctrl+O"), self, self.browse_excel)
        QShortcut(QKeySequence("Ctrl+Return"), self, self.start_process)
        QShortcut(QKeySequence("Ctrl+D"), self, self.browse_pdf_dir)
        QShortcut(QKeySequence("Ctrl+E"), self, self.export_results)
        QShortcut(QKeySequence("Delete"), self.table, self._delete_selected_rows)
        QShortcut(QKeySequence("Ctrl+F"), self, lambda: self.search_input.setFocus())

    def _check_for_resume(self):
        """检查是否有中断的处理会话，提供恢复选项"""
        cp = load_checkpoint()
        if cp is None:
            return
        output_dir, text_settings, tasks, valid_indices = cp
        if not tasks:
            clear_checkpoint()
            return

        elapsed = datetime.now()
        try:
            ts = os.path.getmtime(CHECKPOINT_FILE)
            elapsed = datetime.fromtimestamp(ts)
        except Exception as e:
            diag_log(f"Failed to get checkpoint file modification time: {e}")
        time_str = elapsed.strftime("%Y-%m-%d %H:%M")

        reply = QMessageBox.question(self, "发现中断的任务",
            f"检测到上次未完成的任务（{time_str}）\n"
            f"共 {len(tasks)} 个任务，已处理 {sum(1 for t in tasks if t.get('status') in ('成功','失败'))} 个\n\n"
            "是否恢复上次的处理进度？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes)

        if reply == QMessageBox.StandardButton.Yes:
            self.tasks = tasks
            self.text_settings = text_settings
            if output_dir and os.path.isdir(output_dir):
                self.output_dir.setText(output_dir)
            self.update_style_preview()
            self.update_table()
            self.update_stats()
            self.log(f"已恢复中断会话（{time_str}），{len(self.tasks)} 个任务")
            self._resume_from_checkpoint(valid_indices)
        else:
            clear_checkpoint()
            self.log("已忽略中断会话")

    def _resume_from_checkpoint(self, valid_indices):
        """从检查点恢复处理"""
        ready = [t for t in self.tasks if t['status'] in ('就绪', '未找到') and t.get('pdf_path')]
        if not ready:
            self.log("所有任务已处理完成")
            clear_checkpoint()
            return

        self.log(f"重新处理 {len(ready)} 个未完成的任务...")
        self._start_with_tasks(ready)

    # --- 拖拽支持 ---
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if not urls:
            return
        path = urls[0].toLocalFile()
        if not path:
            return
        if os.path.isfile(path) and path.lower().endswith(('.xlsx', '.xls')):
            self.excel_path.setText(path)
            self.log(f"已拖入 Excel: {os.path.basename(path)}")
        elif os.path.isdir(path):
            pdfs = [f for f in os.listdir(path) if f.lower().endswith('.pdf')]
            if pdfs:
                self.pdf_dir.setText(path)
                self.log(f"已拖入 PDF 文件夹: {os.path.basename(path)} ({len(pdfs)} 个 PDF)")
                if not self.output_dir.text():
                    self.output_dir.setText(os.path.join(path, "output"))
            else:
                self.output_dir.setText(path)
                self.log(f"已拖入输出目录: {os.path.basename(path)}")
        elif os.path.isfile(path) and path.lower().endswith('.pdf'):
            parent = os.path.dirname(path)
            self.pdf_dir.setText(parent)
            self.log(f"已拖入 PDF: {os.path.basename(path)}")

    # --- 表格右键菜单 ---
    def _parse_preview_page_range(self, pr, total_pages):
        return parse_page_range(pr, total_pages)

    def _on_table_context_menu(self, pos):
        row = self.table.rowAt(pos.y())
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background-color: white;
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                padding: 4px;
            }}
            QMenu::item {{
                padding: 8px 24px;
                border-radius: 4px;
                color: {COLORS['text']};
            }}
            QMenu::item:selected {{
                background-color: {COLORS['primary_light']};
                color: {COLORS['primary']};
            }}
        """)

        # 行操作（仅在点击有效行时显示）
        if 0 <= row < len(self.tasks):
            task = self.tasks[row]
            act_edit = menu.addAction("编辑文字")
            if task['pdf_path']:
                act_open = menu.addAction("打开 PDF 文件")
                act_folder = menu.addAction("打开所在文件夹")
            else:
                act_open = act_folder = None
            if task['status'] == '成功' and task['pdf_path']:
                act_open_output = menu.addAction("打开输出文件")
            else:
                act_open_output = None
            menu.addSeparator()
            act_remove = menu.addAction("删除此任务")
            menu.addSeparator()

        # 选择操作（始终可用）
        act_select_all = menu.addAction("全选")
        act_deselect = menu.addAction("取消选择")
        menu.addSeparator()
        # 删除选中行
        selected_rows = set(idx.row() for idx in self.table.selectionModel().selectedRows())
        if selected_rows:
            act_del_selected = menu.addAction(f"删除选中行 ({len(selected_rows)})")
            act_batch_edit = menu.addAction("批量修改选中行文字")
        else:
            act_del_selected = None
            act_batch_edit = None
        # 导出结果
        if self.tasks:
            act_export = menu.addAction("导出结果到 Excel")
        else:
            act_export = None

        action = menu.exec(self.table.viewport().mapToGlobal(pos))

        if action is None:
            return
        if action == act_select_all:
            self.table.selectAll()
            return
        if action == act_deselect:
            self.table.clearSelection()
            return
        if act_del_selected and action == act_del_selected:
            self._delete_selected_rows()
            return
        if act_batch_edit and action == act_batch_edit:
            self._batch_edit_text()
            return
        if act_export and action == act_export:
            self.export_results()
            return
        if 0 <= row < len(self.tasks):
            task = self.tasks[row]
            if action == act_edit:
                item = self.table.item(row, 2)
                if item:
                    self.table.editItem(item)
                return
            if act_open and action == act_open:
                self._open_file(task['pdf_path'])
            elif act_folder and action == act_folder:
                self._open_folder(task['pdf_path'])
            if act_open_output and action == act_open_output:
                out_dir = self.output_dir.text().strip()
                out_file = os.path.join(out_dir, os.path.basename(task['pdf_path']))
                if os.path.exists(out_file):
                    self._open_file(out_file)
                else:
                    self._open_folder(out_dir)
            if action == act_remove:
                self.tasks.pop(row)
                self.update_table()
                self.update_stats()

    def _on_table_double_click(self, row, col):
        """双击表格行：编辑文字列或打开 PDF"""
        if row < 0 or row >= len(self.tasks):
            return
        task = self.tasks[row]
        if col == 2:
            item = self.table.item(row, col)
            if item:
                self.table.editItem(item)
            return
        if task['pdf_path']:
            self._open_file(task['pdf_path'])

    def _on_cell_changed(self, row, col):
        """表格单元格内容变更时同步到 tasks 数据"""
        if col != 2 or row < 0 or row >= len(self.tasks):
            return
        item = self.table.item(row, col)
        if item is None:
            return
        new_text = item.text().strip()
        if new_text != self.tasks[row]['text']:
            self.tasks[row]['text'] = new_text
            if new_text:
                self.log(f"已修改第 {row + 1} 行文字为: {new_text}")
                save_watermark_history(new_text)
            else:
                self.log(f"第 {row + 1} 行文字已清空")
            # 显示智能推荐
            self._show_smart_suggestions(row)

    def _system_open(self, path):
        """用系统默认程序打开文件或目录（跨平台）"""
        try:
            if not path or not os.path.exists(path):
                self.log(f"打开失败: 路径不存在 - {path}")
                return
            if sys.platform == 'win32':
                os.startfile(path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', path], start_new_session=True,
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                subprocess.Popen(['xdg-open', path], start_new_session=True,
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            self.log(f"打开失败: {e}")

    def _open_file(self, path):
        self._system_open(path)

    def _open_folder(self, path):
        self._system_open(os.path.dirname(path))

    def browse_excel(self):
        start_dir = self.settings.value("last_excel_dir", "")
        if not start_dir or not os.path.isdir(start_dir):
            start_dir = os.path.dirname(self.excel_path.text()) if self.excel_path.text() else ""
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", start_dir,
            "Excel 文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        if path:
            self.excel_path.setText(path)
            self.settings.setValue("last_excel_dir", os.path.dirname(path))
            # 自动检测同目录下的 PDF 文件夹
            if not self.pdf_dir.text().strip():
                excel_dir = os.path.dirname(path)
                try:
                    pdfs = [f for f in os.listdir(excel_dir) if f.lower().endswith('.pdf')]
                except OSError:
                    pdfs = []
                if pdfs:
                    self.pdf_dir.setText(excel_dir)
                    self.log(f"自动检测到 PDF 文件夹: {excel_dir} ({len(pdfs)} 个 PDF)")
                else:
                    try:
                        for sub in os.listdir(excel_dir):
                            sub_path = os.path.join(excel_dir, sub)
                            if os.path.isdir(sub_path):
                                try:
                                    sub_pdfs = [f for f in os.listdir(sub_path) if f.lower().endswith('.pdf')]
                                except OSError:
                                    continue
                                if sub_pdfs:
                                    self.pdf_dir.setText(sub_path)
                                    self.log(f"自动检测到 PDF 文件夹: {sub_path} ({len(sub_pdfs)} 个 PDF)")
                                    break
                    except OSError as e:
                        diag_log(f"Failed to scan subdirectories for PDFs: {e}")

    def browse_pdf_dir(self):
        start_dir = self.settings.value("last_pdf_dir", "")
        if not start_dir or not os.path.isdir(start_dir):
            start_dir = self.pdf_dir.text() if self.pdf_dir.text() else ""
        path = QFileDialog.getExistingDirectory(self, "选择 PDF 文件夹", start_dir)
        if path:
            self.pdf_dir.setText(path)
            self.settings.setValue("last_pdf_dir", path)

    def browse_output_dir(self):
        start_dir = self.settings.value("last_output_dir", "")
        if not start_dir or not os.path.isdir(start_dir):
            start_dir = self.output_dir.text() if self.output_dir.text() else ""
        path = QFileDialog.getExistingDirectory(self, "选择输出目录", start_dir)
        if path:
            self.output_dir.setText(path)
            self.settings.setValue("last_output_dir", path)

    def open_output_dir(self):
        path = self.output_dir.text().strip()
        if not path:
            pd = self.pdf_dir.text().strip()
            if pd:
                path = os.path.join(pd, "output")
            else:
                path = os.path.join(os.getcwd(), "output")
        if not os.path.exists(path):
            try:
                os.makedirs(path, exist_ok=True)
            except Exception:
                QMessageBox.warning(self, "提示", f"无法创建输出目录:\n{path}")
                return
        self._system_open(path)

    def preview_pdf(self):
        """预览选中行的 PDF 文字添加效果"""
        if not self.tasks:
            QMessageBox.information(self, "预览", "请先加载 Excel 数据")
            return

        selected = self.table.selectionModel().selectedRows()
        if selected:
            row = selected[0].row()
            if 0 <= row < len(self.tasks) and self.tasks[row]['pdf_path']:
                task = self.tasks[row]
            else:
                QMessageBox.warning(self, "预览", "选中的任务没有关联的 PDF 文件")
                return
        else:
            valid = [t for t in self.tasks if t['pdf_path']]
            if not valid:
                QMessageBox.warning(self, "预览", "没有可预览的 PDF 文件")
                return
            task = valid[0]

        pdf_path = task['pdf_path']
        text = task['text']
        page_texts = task.get('page_texts', [])

        if page_texts:
            _, text = page_texts[0]

        if not text or not text.strip():
            text = task.get('text', '预览文字')

        dialog = PdfPreviewDialog(pdf_path, text, self.text_settings, self)
        dialog.exec()

    def reset_all(self):
        """重置所有设置和任务"""
        reply = QMessageBox.question(self, "确认重置",
            "确定要清空所有设置和任务列表吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return

        self.excel_path.clear()
        self.pdf_dir.clear()
        self.output_dir.clear()
        self.tasks.clear()
        self.update_table()
        self.update_stats()
        self.log_text.clear()
        self.progress.setValue(0)
        self.progress.setFormat("%v/%m")
        self.status_label.setText("就绪")
        _DIR_FILELIST_CACHE.clear()

        self.text_settings = {
            'font_size': DEFAULT_FONT_SIZE,
            'color': DEFAULT_TEXT_COLOR,
            'position': DEFAULT_POSITION,
            'opacity': 1.0,
            'bold': False,
            'italic': False,
            'page_range': '',
            'offset_x': 0,
            'offset_y': 0,
        }
        self.update_style_preview()
        self.log("已重置所有设置")

    def _on_suggestion_selected(self, text):
        """用户从推荐面板选中文字时"""
        current_row = self.table.currentRow()
        if 0 <= current_row < len(self.tasks):
            self.tasks[current_row]['text'] = text
            item = self.table.item(current_row, 2)
            if item:
                item.setText(text)
            self.log(f"已应用推荐文字: {text}")
            save_watermark_history(text)
        self.suggest_widget.hide_suggestions()

    def _show_smart_suggestions(self, row):
        """显示智能推荐文字"""
        if row < 0 or row >= len(self.tasks):
            self.suggest_widget.hide_suggestions()
            return
        filename = self.tasks[row].get('filename', '')
        history = load_watermark_history()
        suggestions = smart_recommend_text(filename, history)
        if suggestions:
            self.suggest_widget.show_suggestions(suggestions, self._on_suggestion_selected)
        else:
            self.suggest_widget.hide_suggestions()

    def one_click_smart_process(self):
        """一键智能处理：扫描PDF→生成任务→自动推荐→处理→报告"""
        pdf_dir = self.pdf_dir.text().strip()
        if not pdf_dir or not os.path.isdir(pdf_dir):
            QMessageBox.warning(self, "提示", "请先选择 PDF 文件夹")
            return

        self.log("=" * 50)
        self.log("🤖 一键智能处理启动...")

        try:
            pdf_files = [f for f in sorted(os.listdir(pdf_dir)) if f.lower().endswith('.pdf')]
        except OSError:
            pdf_files = []
        if not pdf_files:
            QMessageBox.information(self, "提示", "该文件夹下没有 PDF 文件")
            return

        self.log(f"📂 扫描到 {len(pdf_files)} 个 PDF 文件")

        self.tasks = []
        for i, pdf_file in enumerate(pdf_files):
            filename = os.path.splitext(pdf_file)[0]
            pdf_path = os.path.join(pdf_dir, pdf_file)
            history = load_watermark_history()
            suggestions = smart_recommend_text(filename, history)
            default_text = suggestions[0] if suggestions else "处理完成"
            self.tasks.append({
                'row': i + 1, 'filename': filename,
                'text': default_text, 'page_texts': [],
                'pdf_path': pdf_path, 'pages': '',
                'status': '就绪'
            })

        self.update_table()
        self.update_stats()

        self._load_pages_async()

        if not self.output_dir.text().strip():
            self.output_dir.setText(os.path.join(pdf_dir, "output"))

        self.log(f"✅ 已自动生成 {len(self.tasks)} 个任务")

        valid = [t for t in self.tasks if t['pdf_path'] and t['status'] in ('就绪', '失败')]
        if not valid:
            QMessageBox.warning(self, "提示", "没有可处理的 PDF 文件")
            return

        reply = QMessageBox.question(self, "确认一键处理",
            f"将处理 {len(valid)} 个 PDF 文件\n"
            f"输出目录: {self.output_dir.text() or os.path.join(pdf_dir, 'output')}\n\n"
            "是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes)

        if reply == QMessageBox.StandardButton.Yes:
            self.log(f"🚀 开始一键智能处理 {len(valid)} 个文件...")
            self._start_with_tasks(valid)

    def open_settings(self):
        dialog = SettingsDialog(self, self.text_settings)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.text_settings = dialog.get_settings()
            self.update_style_preview()

    def update_style_preview(self):
        s = self.text_settings
        color = s.get('color', DEFAULT_TEXT_COLOR)
        pos = s.get('position', DEFAULT_POSITION)
        size = s.get('font_size', DEFAULT_FONT_SIZE)
        opacity = s.get('opacity', 1.0)
        bold = s.get('bold', False)
        italic = s.get('italic', False)

        font = QFont("", max(10, size // 2))
        font.setBold(bold)
        font.setItalic(italic)
        self.preview_text.setFont(font)
        self.preview_text.setStyleSheet(f"color:{color};border:none;background:transparent;")

        parts = []
        if bold:
            parts.append("粗体")
        if italic:
            parts.append("斜体")
        detail = f"{size}pt"
        if parts:
            detail += f" | {' '.join(parts)}"
        detail += f" | {pos}"
        if opacity < 1.0:
            detail += f" | {int(opacity*100)}%"
        self.style_detail.setText(detail)

    def log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.appendPlainText(f"[{ts}] {message}")
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def load_excel(self):
        excel_path = self.excel_path.text().strip()
        pdf_dir = self.pdf_dir.text().strip()

        if not excel_path:
            QMessageBox.warning(self, "提示", "请先选择 Excel 文件")
            return
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, "错误", "Excel 文件不存在")
            return

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            self.tasks = []
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                QMessageBox.warning(self, "提示", "Excel 文件为空")
                return

            header = rows[0]
            data_rows = rows[1:]
            col_idx = self.find_columns(header)
            self.log(f"Excel 表头识别: {col_idx}")

            if col_idx['filename'] is None or col_idx['text'] is None:
                QMessageBox.warning(self, "表头识别失败",
                    "未能识别'文件名'和'添加文字'列\n请检查表头是否正确")
                return

            max_col = max(v for v in col_idx.values() if v is not None) if any(v is not None for v in col_idx.values()) else -1

            # 第一阶段：按行读取原始数据，并按文件名分组
            raw_entries = []
            for i, row in enumerate(data_rows, 1):
                if not row:
                    continue
                if max_col >= 0 and len(row) <= max_col:
                    continue
                filename = row[col_idx['filename']] if col_idx['filename'] is not None and col_idx['filename'] < len(row) else None
                text = row[col_idx['text']] if col_idx['text'] is not None and col_idx['text'] < len(row) else None
                page_range = row[col_idx['page']] if col_idx.get('page') is not None and col_idx['page'] < len(row) else None
                if filename is None or text is None:
                    continue
                filename = str(filename).strip()
                text = str(text).strip()
                if not filename or not text:
                    continue
                page_range = str(page_range).strip() if page_range is not None else ""
                raw_entries.append({
                    'row': i, 'filename': filename, 'text': text,
                    'page_range': page_range
                })

            # 第二阶段：按文件名合并多行
            grouped = {}
            for entry in raw_entries:
                key = entry['filename']
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(entry)

            for filename, entries in grouped.items():
                pdf_path = self.find_pdf_file(filename, pdf_dir)
                default_entries = [e for e in entries if not e['page_range']]
                page_entries = [e for e in entries if e['page_range']]
                default_text = default_entries[0]['text'] if default_entries else entries[0]['text']
                page_texts = [(e['page_range'], e['text']) for e in page_entries]
                min_row = min(e['row'] for e in entries)
                self.tasks.append({
                    'row': min_row, 'filename': filename,
                    'text': default_text,
                    'page_texts': page_texts,
                    'pdf_path': pdf_path, 'pages': '',
                    'status': '就绪' if pdf_path else '未找到'
                })

            # === 智能冲突/重复检测 ===
            self._detect_conflicts_and_duplicates(pdf_dir)

            self.update_table()
            self.update_stats()
            self.log(f"已加载 {len(self.tasks)} 条任务")

            # 后台加载 PDF 页数
            self._load_pages_async()

            if pdf_dir and not self.output_dir.text():
                self.output_dir.setText(os.path.join(pdf_dir, "output"))

        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取 Excel 失败:\n{str(e)}")
            self.log(f"读取 Excel 失败: {str(e)}")

    def _detect_conflicts_and_duplicates(self, pdf_dir):
        """智能检测：重复文件、同一文件冲突任务、缺失PDF、页码越界"""
        if not self.tasks:
            return

        # 1. 检测同一文件名多次出现但不同文字/页码的冲突
        for task in self.tasks:
            if len(task.get('page_texts', [])) > 1:
                page_map = {}
                for pr, txt in task['page_texts']:
                    indices = parse_page_range(pr, 9999)
                    for idx in indices:
                        if idx in page_map and page_map[idx] != txt:
                            self.log(f"⚠ 冲突: '{task['filename']}' 第{idx+1}页有不同文字: '{page_map[idx]}' vs '{txt}'")
                        page_map[idx] = txt

        # 2. 检测 PDF 文件夹中重复文件
        if pdf_dir and os.path.isdir(pdf_dir):
            try:
                all_pdfs = [f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf')]
                name_count = {}
                for f in all_pdfs:
                    name = os.path.splitext(f)[0]
                    name_count[name] = name_count.get(name, 0) + 1
                for name, cnt in name_count.items():
                    if cnt > 1:
                        self.log(f"⚠ 重复文件名: '{name}' 在文件夹中出现 {cnt} 次")
            except Exception as e:
                diag_log(f"Failed to check for duplicate PDFs: {e}")

        # 3. 检测任务列表中引用的 PDF 是否存在
        missing = [t for t in self.tasks if not t['pdf_path']]
        if missing:
            self.log(f"⚠ {len(missing)} 个任务找不到对应 PDF 文件")
            for t in missing[:5]:
                self.log(f"   - {t['filename']}")
            if len(missing) > 5:
                self.log(f"   ... 共 {len(missing)} 个")

        # 4. 智能推荐：对于未找到 PDF 的任务，尝试模糊匹配建议
        for task in self.tasks:
            if not task['pdf_path'] and pdf_dir and os.path.isdir(pdf_dir):
                try:
                    files = os.listdir(pdf_dir)
                    name = task['filename'].lower()
                    if name.endswith('.pdf'):
                        name = name[:-4]
                    match, score = self._smart_match_pdf(name, pdf_dir, files)
                    if match and score >= 60:
                        self.log(f"💡 建议: '{task['filename']}' 可能匹配 '{os.path.basename(match)}' (置信度{score}%)")
                except Exception as e:
                    diag_log(f"Failed to smart match PDF: {e}")

    def _load_pages_async(self):
        """后台加载 PDF 页数"""
        if not self.tasks:
            return
        self._page_worker = PageCountWorker(self.tasks, self)
        self._page_worker.page_loaded.connect(self._on_page_loaded)
        self._page_worker.start()

    def _on_page_loaded(self, index, pages):
        """单个 PDF 页数加载完成，更新数据模型和表格"""
        if 0 <= index < len(self.tasks):
            self.tasks[index]['pages'] = pages
        if 0 <= index < self.table.rowCount():
            item = self.table.item(index, 4)
            if item:
                item.setText(pages)

    def find_columns(self, header):
        col_idx = {'no': None, 'filename': None, 'text': None, 'page': None}
        for i, col in enumerate(header):
            if col is None:
                continue
            cs = str(col).strip()
            if cs in ['序号', '编号', 'ID', 'No', 'no']:
                col_idx['no'] = i
            if cs in ['文件名', '文件', 'PDF', 'pdf', '名称', '文件名称']:
                col_idx['filename'] = i
            if cs in ['添加的文字内容', '文字内容', '文字', '内容', '添加内容', '添加文字', '水印文字', '标注']:
                col_idx['text'] = i
            if cs in ['页码', '页码范围', '页面', '页', 'Page', 'page', 'Pages', 'pages']:
                col_idx['page'] = i
        if col_idx['filename'] is None and len(header) > 1:
            col_idx['filename'] = 1
        if col_idx['text'] is None and len(header) > 2:
            col_idx['text'] = 2
        return col_idx

    def _list_dir_cached(self, dir_path):
        """缓存目录列表，减少重复 os.listdir"""
        if dir_path not in _DIR_FILELIST_CACHE:
            try:
                _DIR_FILELIST_CACHE[dir_path] = os.listdir(dir_path)
            except OSError:
                _DIR_FILELIST_CACHE[dir_path] = []
        return _DIR_FILELIST_CACHE[dir_path]

    def _smart_match_pdf(self, name, pdf_dir, files):
        """智能模糊匹配PDF文件名，返回 (最佳文件名, 置信度0-100)"""
        name_lower = name.lower()
        # 阶段1: 精确匹配
        for f in files:
            fl = f.lower()
            fn = fl[:-4] if fl.endswith('.pdf') else fl
            if fl == f"{name_lower}.pdf" or fn == name_lower:
                return os.path.join(pdf_dir, f), 100
        # 阶段2: 包含匹配 + SequenceMatcher
        best_score = 0
        best_file = None
        for f in files:
            fl = f.lower()
            if not fl.endswith('.pdf'):
                continue
            fn = fl[:-4]
            score = int(difflib.SequenceMatcher(None, name_lower, fn).ratio() * 100)
            if score > best_score:
                best_score = score
                best_file = f
            if name_lower in fn or fn in name_lower:
                contain_score = 90 + min(len(name_lower), len(fn)) // 10
                if contain_score > best_score:
                    best_score = contain_score
                    best_file = f
        # 阶段3: 去除特殊字符后匹配
        if best_score < 70:
            def normalize(s):
                s = re.sub(r'[\s_\-（）()：:，,．.]', '', s)
                return s
            norm_name = normalize(name_lower)
            for f in files:
                fl = f.lower()
                if not fl.endswith('.pdf'):
                    continue
                fn = fl[:-4]
                if norm_name == normalize(fn):
                    if 85 > best_score:
                        best_score = 85
                        best_file = f
        return os.path.join(pdf_dir, best_file) if best_file else None, best_score

    def find_pdf_file(self, filename, pdf_dir):
        """查找PDF文件路径"""
        if not pdf_dir or not os.path.isdir(pdf_dir):
            return ""

        # 先精确匹配
        candidates = [filename, filename + ".pdf"]
        for c in candidates:
            path = os.path.join(pdf_dir, c)
            if os.path.isfile(path):
                return path

        # 智能模糊匹配
        files = self._list_dir_cached(pdf_dir)
        name = os.path.splitext(filename)[0]
        match, score = self._smart_match_pdf(name, pdf_dir, files)
        if match and score >= 60:
            return match

        return ""

    def update_table(self):
        """更新任务表格"""
        self.table.setRowCount(len(self.tasks))
        for i, task in enumerate(self.tasks):
            self.table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.table.setItem(i, 1, QTableWidgetItem(task.get('filename', '')))
            text = task.get('text', '')
            self.table.setItem(i, 2, QTableWidgetItem(text))
            page_range = task.get('page_range', '')
            if task.get('page_texts'):
                page_range = ", ".join(f"{pr}" for pr, _ in task['page_texts'])
            self.table.setItem(i, 3, QTableWidgetItem(page_range))
            self.table.setItem(i, 4, QTableWidgetItem(str(task.get('pages', ''))))
            self.table.setItem(i, 5, QTableWidgetItem(task.get('pdf_path', '') or "未找到"))

            status = task.get('status', '就绪')
            status_item = QTableWidgetItem(status)
            self.table.item(i, 4)

        self.task_count.setText(f"{len(self.tasks)} 个任务")

    def update_stats(self):
        """更新统计信息"""
        total = len(self.tasks)
        ready = sum(1 for t in self.tasks if t.get('status') == '就绪')
        missing = sum(1 for t in self.tasks if t.get('status') == '未找到')
        self.stat_total.setText(f"总任务 {total}")
        self.stat_ready.setText(f"就绪 {ready}")
        self.stat_missing.setText(f"未找到 {missing}")

    def _on_header_sort(self, column):
        """表头点击排序"""
        if column == self._sort_col:
            self._sort_order = Qt.SortOrder.DescendingOrder if self._sort_order == Qt.SortOrder.AscendingOrder else Qt.SortOrder.AscendingOrder
        else:
            self._sort_col = column
            self._sort_order = Qt.SortOrder.AscendingOrder

        # 排序 tasks
        col_name = ['no', 'filename', 'text', 'page_range', 'pages', 'pdf_path', 'status']
        key_name = col_name[column] if column < len(col_name) else 'filename'
        self.tasks.sort(
            key=lambda t: (t.get(key_name, '') or '') if isinstance(t.get(key_name), str) else str(t.get(key_name, '') or ''),
            reverse=(self._sort_order == Qt.SortOrder.DescendingOrder)
        )
        self.update_table()

    def _on_search_changed(self):
        """搜索框内容变化时防抖"""
        self._search_timer.start()

    def _do_filter_table(self):
        """执行表格过滤"""
        query = self.search_input.text().lower()
        for i, task in enumerate(self.tasks):
            row = self.table.row(i)
            filename = (task.get('filename', '') or '').lower()
            text = (task.get('text', '') or '').lower()
            if query and query not in filename and query not in text:
                self.table.setRowHidden(i, True)
            else:
                self.table.setRowHidden(i, False)

    def _open_batch_edit(self):
        """打开批量编辑对话框"""
        if not self.tasks:
            QMessageBox.warning(self, "提示", "请先加载任务列表")
            return
        dialog = BatchEditDialog(self.tasks, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            tasks = dialog.get_results()
            self.tasks = tasks
            self.update_table()

    def _run_preflight_check(self):
        """运行预检对话框"""
        if not self.tasks:
            QMessageBox.warning(self, "提示", "请先加载任务列表")
            return
        dialog = PreflightDialog(self.tasks, self)
        dialog.exec()

    def _open_template_manager(self):
        """打开水印模板管理对话框"""
        dialog = WatermarkTemplateDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            template = dialog.get_selected_template()
            if template:
                self.text_settings = {
                    'font_size': template.get('font_size', DEFAULT_FONT_SIZE),
                    'color': template.get('color', DEFAULT_TEXT_COLOR),
                    'position': template.get('position', DEFAULT_POSITION),
                    'opacity': template.get('opacity', 1.0),
                    'bold': template.get('bold', False),
                    'italic': template.get('italic', False),
                    'page_range': template.get('page_range', ''),
                    'offset_x': template.get('offset_x', 0),
                    'offset_y': template.get('offset_y', 0),
                }
                self.update_style_preview()
                self.log(f"已应用水印模板: {template.get('name', '')}")

    def _open_pdf_tools(self):
        """打开 PDF 辅助工具对话框（页码/页脚/文本提取）"""
        dialog = PdfToolsDialog(self)
        dialog.exec()

    def _open_advanced_tools(self):
        """打开 PDF 增强工具对话框（图片水印/旋转/加密）"""
        dialog = PDFAdvancedDialog(self)
        dialog.exec()

    def _open_multi_text(self):
        """打开批量多文字配置对话框"""
        dialog = MultiTextDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            multi_texts = dialog.get_texts()
            if not multi_texts:
                return
            # 应用到所有已加载的任务
            count = 0
            for task in self.tasks:
                # 只有已加载 PDF 的任务才应用多文字
                if task.get('pdf_path') and os.path.exists(task['pdf_path']):
                    task['text'] = multi_texts
                    count += 1
            if count > 0:
                self.log(f"批量多文字: 已应用 {len(multi_texts)} 个文字条目到 {count} 个任务")
                self.update_table()

    def _smart_classify_tasks(self):
        """智能分类任务"""
        if not self.tasks:
            QMessageBox.warning(self, "提示", "请先加载任务列表")
            return
        for task in self.tasks:
            filename = task.get('filename', '').lower()
            if '合同' in filename or '协议' in filename:
                task['text'] = '已审核'
            elif '发票' in filename or '收据' in filename:
                task['text'] = '已报销'
            elif '草稿' in filename:
                task['text'] = '草稿'
            elif '最终' in filename or '终版' in filename:
                task['text'] = '最终版'
        self.update_table()
        self.log("📂 已根据文件名智能分类任务")

    def _restore_from_backup(self):
        """还原原始 PDF 文件"""
        QMessageBox.information(self, "提示", "还原功能需要备份目录，请手动操作")

    def _smart_sync_watermark(self):
        """同步水印"""
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "提示", "请先选中要同步的水印来源")
            return
        source_text = self.tasks[selected[0].row()].get('text', '')
        for task in self.tasks:
            task['text'] = source_text
        self.update_table()
        self.log(f"🔗 已同步水印文字: {source_text}")

    def _adaptive_page_size(self):
        """自适应页面大小"""
        if not self.tasks:
            QMessageBox.warning(self, "提示", "请先加载任务列表")
            return
        QMessageBox.information(self, "提示", "页面适应功能需要分析 PDF 页面尺寸，请手动设置")

    def _delete_selected_rows(self):
        """删除选中行"""
        selected_rows = sorted(set(idx.row() for idx in self.table.selectionModel().selectedRows()), reverse=True)
        if not selected_rows:
            return
        for row in selected_rows:
            if 0 <= row < len(self.tasks):
                self.tasks.pop(row)
        self.update_table()
        self.update_stats()
        self.log(f"已删除 {len(selected_rows)} 个任务")

    def _batch_edit_text(self):
        """批量修改选中行文字"""
        self._open_batch_edit()

    def export_results(self):
        """导出结果到 Excel"""
        if not self.tasks:
            QMessageBox.warning(self, "提示", "没有任务可导出")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "导出结果", os.path.join(os.getcwd(), "处理结果.xlsx"),
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "处理结果"
            ws.append(['序号', '文件名', '添加文字', '页码', '页数', 'PDF路径', '状态'])
            for i, task in enumerate(self.tasks, 1):
                ws.append([
                    i, task.get('filename', ''), task.get('text', ''),
                    ', '.join(f"{pr}" for pr, _ in task.get('page_texts', [])),
                    task.get('pages', ''), task.get('pdf_path', ''), task.get('status', '')
                ])
            wb.save(path)
            wb.close()
            self.log(f"已导出结果: {path}")
            QMessageBox.information(self, "成功", f"已导出 {len(self.tasks)} 条结果到:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def download_template(self):
        """下载 Excel 模板"""
        path, _ = QFileDialog.getSaveFileName(
            self, "保存模板", os.path.join(os.getcwd(), "PDF批量添加文字模板.xlsx"),
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "任务列表"
            ws.append(['序号', '文件名', '添加文字', '页码范围'])
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            wb.save(path)
            wb.close()
            self.log(f"已生成模板: {path}")
            QMessageBox.information(self, "成功", f"模板已保存到:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成模板失败:\n{str(e)}")

    def auto_generate_excel(self):
        """智能生成 Excel 任务表"""
        pdf_dir = self.pdf_dir.text().strip()
        if not pdf_dir or not os.path.isdir(pdf_dir):
            QMessageBox.warning(self, "提示", "请先选择 PDF 文件夹")
            return

        try:
            pdf_files = sorted(f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf'))
            if not pdf_files:
                QMessageBox.information(self, "提示", "该文件夹下没有 PDF 文件")
                return

            path, _ = QFileDialog.getSaveFileName(
                self, "保存任务表", os.path.join(pdf_dir, "任务表.xlsx"),
                "Excel 文件 (*.xlsx)"
            )
            if not path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "任务列表"
            ws.append(['序号', '文件名', '添加文字', '页码范围'])

            for i, pdf_file in enumerate(pdf_files, 1):
                filename = os.path.splitext(pdf_file)[0]
                history = load_watermark_history()
                suggestions = smart_recommend_text(filename, history)
                suggested_text = suggestions[0] if suggestions else "处理完成"
                ws.append([i, filename, suggested_text, ""])

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            wb.save(path)
            wb.close()
            QMessageBox.information(self, "成功", f"已为 {len(pdf_files)} 个 PDF 自动生成任务表:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成失败:\n{str(e)}")

    def clear_tasks(self):
        """清空任务列表"""
        self.tasks.clear()
        self.update_table()
        self.update_stats()
        self.log("已清空任务列表")

    def start_process(self):
        """开始处理"""
        if not self.tasks:
            QMessageBox.warning(self, "提示", "请先加载任务列表")
            return

        valid_tasks = [t for t in self.tasks if t.get('pdf_path')]
        if not valid_tasks:
            QMessageBox.warning(self, "提示", "没有有效的 PDF 文件可处理")
            return

        reply = QMessageBox.question(self, "确认处理",
            f"将处理 {len(valid_tasks)} 个 PDF 文件\n"
            f"输出目录: {self.output_dir.text() or os.path.join(self.pdf_dir.text(), 'output')}\n\n"
            "是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes)

        if reply != QMessageBox.StandardButton.Yes:
            return

        self._start_with_tasks(valid_tasks)

    def _start_with_tasks(self, tasks):
        """启动处理线程"""
        self.worker = ProcessWorker(
            tasks=tasks,
            output_dir=self.output_dir.text().strip(),
            font_size=self.text_settings.get('font_size', DEFAULT_FONT_SIZE),
            text_color=self.text_settings.get('color', DEFAULT_TEXT_COLOR),
            position=self.text_settings.get('position', DEFAULT_POSITION),
            page_range=self.text_settings.get('page_range', ''),
            opacity=self.text_settings.get('opacity', 1.0),
            bold=self.text_settings.get('bold', False),
            italic=self.text_settings.get('italic', False),
            offset_x=self.text_settings.get('offset_x', 0),
            offset_y=self.text_settings.get('offset_y', 0),
            naming_template=self.naming_template.currentText(),
        )

        self.worker.progress.connect(self._on_progress)
        self.worker.status.connect(self.status_label.setText)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.task_status.connect(self._on_task_status)
        self.worker.log.connect(self.log)

        # 保存检查点
        save_checkpoint(
            self.output_dir.text().strip(),
            self.text_settings,
            self.tasks,
            [i for i, t in enumerate(self.tasks) if t in tasks]
        )

        self._start_time = datetime.now()
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setValue(0)
        self.status_label.setText("处理中...")
        self.log(f"开始处理 {len(tasks)} 个 PDF 文件")
        self.worker.start()

    def stop_process(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("用户请求停止处理")

    def _on_progress(self, current, total):
        """处理进度更新"""
        self.progress.setMaximum(total)
        self.progress.setValue(current)
        self.progress.setFormat(f"{current}/{total}")

    def _on_task_status(self, index, status_text):
        """单个任务状态更新"""
        if 0 <= index < len(self.tasks):
            self.tasks[index]['status'] = status_text
            item = self.table.item(index, 6)
            if item:
                item.setText(status_text)

    def on_finished(self, success, msg):
        """处理完成"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        # 清除检查点
        all_done = all(t.get('status') in ('成功', '失败', '跳过') for t in self.tasks)
        if all_done:
            clear_checkpoint()

        elapsed = ""
        if hasattr(self, '_start_time'):
            delta = datetime.now() - self._start_time
            seconds = int(delta.total_seconds())
            if seconds >= 60:
                elapsed = f" (耗时 {seconds // 60}分{seconds % 60}秒)"
            else:
                elapsed = f" (耗时 {seconds}秒)"

        self.status_label.setText(msg + elapsed)
        self.log("=" * 40)
        self.log(msg + elapsed)
        self.update_stats()

        has_failed = any(t['status'] == '失败' for t in self.tasks)
        self.retry_btn.setVisible(has_failed)

        has_results = any(t.get('status') in ('成功', '失败') for t in self.tasks)
        self.report_btn.setVisible(has_results)
        if has_results:
            for t in self.tasks:
                if t.get('text') and t['text'].strip():
                    save_watermark_history(t['text'])

        if success:
            send_system_notification("PDF批量添加文字", "处理完成！所有文件已成功添加文字")
        else:
            send_system_notification("PDF批量添加文字", "处理完成，部分文件处理失败")

        output_paths = getattr(self.worker, 'output_paths', [])
        out_path = getattr(self.worker, 'output_dir', None) or self.output_dir.text().strip()

        if success:
            file_list = ""
            if output_paths:
                file_list = "\n\n输出文件:\n" + "\n".join(f"  - {os.path.basename(p)}" for p in output_paths[:10])
                if len(output_paths) > 10:
                    file_list += f"\n  ... 共 {len(output_paths)} 个文件"
                actual_out_dir = os.path.dirname(output_paths[0])
                if actual_out_dir != out_path:
                    out_path = actual_out_dir
                    self.output_dir.setText(out_path)
            elif os.path.exists(out_path):
                files = [f for f in os.listdir(out_path) if f.lower().endswith('.pdf')]
                if files:
                    file_list = "\n\n输出文件:\n" + "\n".join(f"  - {f}" for f in sorted(files)[:10])
                    if len(files) > 10:
                        file_list += f"\n  ... 共 {len(files)} 个文件"
            detail = f"{msg}{file_list}\n\n输出目录:\n{out_path}"
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("处理完成")
            msg_box.setText(detail)
            msg_box.setIcon(QMessageBox.Icon.Information)
            open_btn = msg_box.addButton("打开输出目录", QMessageBox.ButtonRole.AcceptRole)
            close_btn = msg_box.addButton("关闭", QMessageBox.ButtonRole.RejectRole)
            msg_box.exec()
            if msg_box.clickedButton() == open_btn:
                self._system_open(out_path)
        else:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("处理完成")
            msg_box.setText(msg)
            msg_box.setIcon(QMessageBox.Icon.Warning)
            open_btn = msg_box.addButton("打开输出目录", QMessageBox.ButtonRole.AcceptRole)
            close_btn = msg_box.addButton("关闭", QMessageBox.ButtonRole.RejectRole)
            msg_box.exec()
            if msg_box.clickedButton() == open_btn:
                self._system_open(out_path)

    def retry_failed(self):
        """重试失败项"""
        failed = [t for t in self.tasks if t.get('status') == '失败']
        if not failed:
            return
        self._start_with_tasks(failed)

    def generate_report(self):
        """生成处理报告"""
        if not self.tasks:
            return
        success = sum(1 for t in self.tasks if t.get('status') == '成功')
        failed = sum(1 for t in self.tasks if t.get('status') == '失败')
        skipped = sum(1 for t in self.tasks if t.get('status') == '跳过')
        total = len(self.tasks)
        report = (
            f"=== 处理报告 ===\n"
            f"总任务: {total}\n"
            f"成功: {success}\n"
            f"失败: {failed}\n"
            f"跳过: {skipped}\n"
        )
        self.log(report)
        QMessageBox.information(self, "处理报告", report)

    def load_settings(self):
        """加载上次保存的设置"""
        geo = self.settings.value("geometry")
        if geo:
            self.restoreGeometry(geo)

        excel_val = self.settings.value("excel_path", "")
        if excel_val and os.path.isfile(excel_val):
            self.excel_path.setText(excel_val)
        pdf_val = self.settings.value("pdf_dir", "")
        if pdf_val and os.path.isdir(pdf_val):
            self.pdf_dir.setText(pdf_val)
        out_val = self.settings.value("output_dir", "")
        if out_val and os.path.isdir(out_val):
            self.output_dir.setText(out_val)

        def tb(v):
            if isinstance(v, bool):
                return v
            if v is None:
                return False
            return str(v).lower() == "true"

        def safe_int(v, default):
            try:
                return int(v)
            except (ValueError, TypeError):
                return default

        def safe_float(v, default):
            try:
                return float(v)
            except (ValueError, TypeError):
                return default

        self.text_settings = {
            'font_size': safe_int(self.settings.value("font_size", DEFAULT_FONT_SIZE), DEFAULT_FONT_SIZE),
            'color': self.settings.value("color", DEFAULT_TEXT_COLOR) or DEFAULT_TEXT_COLOR,
            'position': self.settings.value("position", DEFAULT_POSITION) or DEFAULT_POSITION,
            'opacity': safe_float(self.settings.value("opacity", 1.0), 1.0),
            'bold': tb(self.settings.value("bold", False)),
            'italic': tb(self.settings.value("italic", False)),
            'page_range': self.settings.value("page_range", "") or "",
            'offset_x': safe_int(self.settings.value("offset_x", 0), 0),
            'offset_y': safe_int(self.settings.value("offset_y", 0), 0),
        }
        self.update_style_preview()

    def _check_update_on_startup(self):
        """启动时在后台静默检查更新（安全：不弹窗，只更新状态栏）"""
        from .utils.auto_update import check_for_update
        import threading
        from PyQt6.QtCore import QTimer

        def _check():
            try:
                result = check_for_update()
                QTimer.singleShot(0, lambda: self._on_startup_update_result(result))
            except Exception:
                pass

        threading.Thread(target=_check, daemon=True).start()

    def _on_startup_update_result(self, result):
        """主线程：根据启动时检查结果更新状态栏"""
        if result.get("has_update") and result.get("latest_version"):
            self.status_label.setText(
                f"发现新版本 v{result['latest_version']} | 点击版本号下载"
            )
            self.status_label.setStyleSheet(
                f"color:{COLORS['accent']};font-size:13px;font-weight:700;"
                f"border:1px solid {COLORS['accent']};border-radius:8px;"
                f"padding:4px 16px;background:{COLORS['accent_light']};"
            )

    def _show_update_notification(self, latest_version, changelog, download_url):
        """在状态栏显示更新通知"""
        self.status_label.setText(
            f"发现新版本 v{latest_version} | 点击 🔄 检查更新 下载"
        )
        self.status_label.setStyleSheet(
            f"color:{COLORS['accent']};font-size:12px;font-weight:600;border:none;background:transparent;"
        )

    def _check_update_now(self):
        """点击版本号检测更新（后台线程检查，主线程弹窗）"""
        from PyQt6.QtCore import QTimer
        from .utils.auto_update import check_for_update
        import threading

        self.status_label.setText("正在检查更新...")
        self.status_label.setCursor(QCursor(Qt.CursorShape.WaitCursor))

        def _check():
            """后台线程：只做 API 请求"""
            try:
                result = check_for_update()
            except Exception:
                result = {"has_update": False, "latest_version": APP_VERSION,
                          "changelog": "", "download_url": "", "cached": False}
            # 用 QTimer 切回主线程处理弹窗
            QTimer.singleShot(0, lambda: self._on_update_result(result))

        thread = threading.Thread(target=_check, daemon=True)
        thread.start()

    def _on_update_result(self, result):
        """主线程处理更新结果（安全弹窗）"""
        from PyQt6.QtWidgets import QMessageBox
        import webbrowser

        self.status_label.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        has_update = result.get("has_update", False)
        latest_version = result.get("latest_version", "")
        changelog = result.get("changelog", "")
        download_url = result.get("download_url", "")

        if has_update and latest_version:
            self._show_update_notification(latest_version, changelog, download_url)
            msg = QMessageBox(self)
            msg.setWindowTitle("发现新版本")
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setText(
                f"<b>新版本 v{latest_version}</b> 可用！<br>"
                f"当前版本：v{APP_VERSION}"
            )
            if changelog:
                preview = changelog[:300] + ("..." if len(changelog) > 300 else "")
                msg.setInformativeText(f"更新内容：\n{preview}")
            download_btn = msg.addButton("下载最新版本", QMessageBox.ButtonRole.AcceptRole)
            later_btn = msg.addButton("稍后提醒", QMessageBox.ButtonRole.RejectRole)
            msg.setDefaultButton(download_btn)
            msg.exec()
            if msg.clickedButton() == download_btn:
                if download_url:
                    webbrowser.open(download_url)
                    self.status_label.setText("已打开下载页面")
                else:
                    self.status_label.setText(f"就绪 | {APP_NAME} v{APP_VERSION}")
        else:
            self.status_label.setText(f"就绪 | {APP_NAME} v{APP_VERSION} (已是最新)")
            QMessageBox.information(self, "检查更新",
                f"当前版本：v{APP_VERSION}\n已是最新版本，无需更新。")

    def closeEvent(self, event):
        self.settings.setValue("geometry", self.saveGeometry())
        self.settings.setValue("excel_path", self.excel_path.text())
        self.settings.setValue("pdf_dir", self.pdf_dir.text())
        self.settings.setValue("output_dir", self.output_dir.text())
        self.settings.setValue("naming_template", self.naming_template.currentText())
        s = self.text_settings
        for k in ['font_size', 'color', 'position', 'opacity', 'bold', 'italic', 'page_range', 'offset_x', 'offset_y']:
            self.settings.setValue(k, s.get(k, ''))
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait(3000)
        if hasattr(self, '_page_worker') and self._page_worker.isRunning():
            self._page_worker.quit()
            self._page_worker.wait(2000)
        event.accept()